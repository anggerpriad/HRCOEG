import openpyxl
import json

xlsx_file = r'Talent Pool #1.xlsx'

try:
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    print("Sheet names:", wb.sheetnames)
    
    data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = row
            else:
                if any(cell is not None for cell in row):
                    row_dict = {headers[j]: row[j] for j in range(len(headers)) if j < len(row)}
                    rows.append(row_dict)
        
        data[sheet_name] = {
            'headers': headers,
            'rows': rows
        }
    
    with open('excel_data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, default=str)
    
    print(f"Data extracted successfully to excel_data.json")
    for sheet, content in data.items():
        print(f"\n{sheet}: {len(content['rows'])} rows")
        
except Exception as e:
    print(f"Error: {e}")
