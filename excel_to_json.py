#!/usr/bin/env python3
"""
Excel to JSON Converter for HRCOEG
Converts Excel file to JSON format for local use if CDN-based XLSX library fails
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed")
    print("Install it with: pip install openpyxl")
    sys.exit(1)

def excel_to_json(excel_file, output_json):
    """Convert Excel file to JSON"""
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            headers = None
            
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                # Skip empty rows
                if not any(cell is not None for cell in row):
                    continue
                
                if i == 0:
                    headers = [str(h) if h is not None else f"Column{j}" for j, h in enumerate(row)]
                else:
                    if headers:
                        row_dict = {}
                        for j, value in enumerate(row):
                            if j < len(headers):
                                row_dict[headers[j]] = value
                        rows.append(row_dict)
            
            if rows:  # Only add sheet if it has data
                data[sheet_name] = {
                    'headers': headers,
                    'rows': rows,
                    'count': len(rows)
                }
        
        # Write to JSON
        with open(output_json, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)
        
        print(f"✅ Successfully converted {excel_file} to {output_json}")
        for sheet, content in data.items():
            print(f"   - {sheet}: {content['count']} rows")
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

if __name__ == '__main__':
    # Convert Talent Pool #1.xlsx
    excel_file = 'Talent Pool #1.xlsx'
    output_file = 'talent-pool-data.json'
    
    if Path(excel_file).exists():
        excel_to_json(excel_file, output_file)
    else:
        print(f"Error: {excel_file} not found in current directory")
        print(f"Current directory: {Path.cwd()}")
